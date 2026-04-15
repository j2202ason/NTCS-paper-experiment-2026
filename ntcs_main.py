import os, datetime, time, ast, gc, json
from langchain_deepseek import ChatDeepSeek
from langchain_core.prompts import ChatPromptTemplate, MessagesPlaceholder
from langchain_core.runnables.history import RunnableWithMessageHistory
from langchain_community.chat_message_histories import ChatMessageHistory
from openpyxl import Workbook, load_workbook
from clscaseset import DataExperiments
from clsprompt import clsprompts
from RAG import *

dskey=''

class MemoryManager:
    def __init__(self):
        self.memory = {}
        self.memory_id = "memory_of_conversation"

    def rec_history(self) -> ChatMessageHistory:
        if self.memory_id not in self.memory:
            self.memory[self.memory_id] = ChatMessageHistory()
        return self.memory[self.memory_id]

    def set_memoryid(self, memory_id): self.memory_id = memory_id

    def get_memory_id(self): return self.memory_id


class BaseReasoner:
    def __init__(self, temperature: int = 0, max_tokens: int = 4096, timeout: int = 180, max_retries: int = 1):  #
        self.llm = ChatDeepSeek(
            model="deepseek-reasoner",  # "deepseek-chat",
            temperature=temperature,
            max_tokens=max_tokens,
            timeout=timeout,
            max_retries=max_retries,
            api_key=os.getenv(dskey), #("DEEPSEEK_API_KEY"),
            api_base="https://api.deepseek.com"
        )
        self._text = ""
        self.chain = None
        self.chain_with_memory = None
        self.memory = MemoryManager
        self.systemMessage = ''
        self.prompt = None

    def set_text(self, dctPrompt: dict):
        self._text = dctPrompt.get("txtHuman")
        self.systemMessage = dctPrompt.get("txtSystem")

    def parser(self,temperature=0):
        self.prompt = ChatPromptTemplate.from_messages([
            ("system", self.systemMessage),
            MessagesPlaceholder(variable_name="history"),
            ("human", "{human_text}")
        ])
        self.llm.temperature=temperature
        self.chain = self.prompt | self.llm

        self.chain_with_memory = RunnableWithMessageHistory(
            runnable=self.chain,
            get_session_history=self.memory.rec_history,
            input_messages_key="human_text",
            history_messages_key="history"
        )

        results = self.chain_with_memory.invoke(
            {"human_text": self._text},
            config={"configurable": {"session_id": self.memory.get_memory_id}}
        )

        return results



class clsExperiments:
    def __init__(self):
        self.prompt = clsprompts
        self.expdata = DataExperiments
        self.currentData = ''
        self.dataid = 0
        self.advmaxtimes = 0
        self.sleep = 0
        self.reasoner = BaseReasoner
        self.exp_id = ''
        self.k_collector = Searching
        self.required_keys = ["exp_id", "case_id", "struct_txt", "cause", "phyContradiction", "causal_chain",
                              "conditions_UNDE", "conditions_DE", "solution_strategies", "solutions"]
        self.exp_dict = {}

    def experimentBuilder(self):
        # for key in self.exp_dict.keys():
        # self.exp_dict[key]=None
        # self.exp_dict['solutions']=[]
        self.exp_dict = dict.fromkeys(self.required_keys)
        # 生成实验序列码
        self.exp_id = datetime.datetime.now().strftime("%Y%m%d%M%S")
        self.exp_dict["exp_id"] = self.exp_id

        # 开始实验，并记录实验数据
        data_id = self.dataid  # int(input('input the id of case(0-3):'))
        advcount = self.advmaxtimes  # int(input('input the max times of experiments(1-200):'))
        self.exp_dict["case_id"] = "Case " + str(self.dataid)
        print(f"【 Case: {self.dataid} 】")
        #############################stage0_BEGIN#########################################
        print("【Checking the input text】")
        stage0_turns = 0
        isaccepted = False
        addinfo = ''
        while stage0_turns < 5 and not isaccepted:
            situations = self.getsituation(data_id, 0, addinfo)
            self.reasoner.set_text(situations)
            results = self.reasoner.parser()
            stage0_turns += 1
            if 'Accept text' in results.content:
                isaccepted = True
            else:
                #addinfo=input(results.content+'\n  Please provide a clearer explanation ：')  # for single experiment
                addinfo = "Proper noun, can be used directly"  # for multiple experiments
        print(results.content)
        ######################################stage0_END################################

        if isaccepted:
            ######################################stage1-2_BEGIN##############################
            print(">>>>Basic analyzing...")
            situations = self.getsituation(data_id, 1, '')
            self.reasoner.set_text(situations)
            results = self.reasoner.parser()
            self.exp_dict["struct_txt"] = results.content
            print(results.content)

            print(">>>>Causal analyzing...")
            situations = self.getsituation(data_id, 2, '')
            self.reasoner.set_text(situations)
            results = self.reasoner.parser()
            self.exp_dict["cause"] = results.content
            print(results.content)
            print(">>>>Finding key attribute...")
            situations = self.getsituation(data_id, 3, '')
            self.reasoner.set_text(situations)
            results = self.reasoner.parser()
            print(results.content)
            print(">>>>Describe the physical contradiction...")
            situations = self.getsituation(data_id, 4, '')
            self.reasoner.set_text(situations)
            results = self.reasoner.parser()
            print(results.content)
            self.exp_dict["phyContradiction"] = results.content

            print(">>>>Expressing the causal chain...")
            situations = self.getsituation(data_id, 5, '')  # chsrootconflict)
            self.reasoner.set_text(situations)
            results = self.reasoner.parser()
            print(results.content)
            self.exp_dict["causal_chain"] = results.content

            print(">>>>Condition analyzing in condition/causal relationship...1...")
            situations = self.getsituation(data_id, 6, '')
            self.reasoner.set_text(situations)
            results = self.reasoner.parser()
            print(results.content)
            self.exp_dict["conditions_UNDE"] = results.content
            print(">>>>Condition analyzing in condition/causal relationship...2...")
            situations = self.getsituation(data_id, 7, '')
            self.reasoner.set_text(situations)
            results = self.reasoner.parser()
            print(results.content)
            self.exp_dict["conditions_DE"] = results.content

            print(">>>>> Solution strategies analyzing...")
            situations = self.getsituation(data_id, 8, '')
            self.reasoner.set_text(situations)
            results = self.reasoner.parser()
            print(results.content)
            self.exp_dict["solution_strategies"] = results.content

            print(">>>>> Recommended solutions are as follows....")
            situations = self.getsituation(data_id, 9, '')
            self.reasoner.set_text(situations)
            results = self.reasoner.parser(1.2)
            print(results.content)
            # jsolution=json.loads(results.content)
            self.exp_dict["solutions"] = results.content


            ############################

    def getsituation(self, data_id, task_id, addinfo):

        if task_id in [25, 26]:
            sysmsg = self.prompt.get_sysprompts(1)
        else:
            sysmsg = self.prompt.get_sysprompts(0)

        if task_id == 0:
            humanmsg = self.expdata.getdata(data_id) + '\n' + addinfo + '\n' + self.prompt.get_humanprompts(task_id)
        elif task_id == 5:
            humanmsg = self.prompt.get_humanprompts(task_id) + '\n' + f"'rootconfilct':{addinfo}"
        else:
            humanmsg = self.prompt.get_humanprompts(task_id)
        return {"txtHuman": humanmsg, "txtSystem": sysmsg}


class Searching:
    def __init__(self):
        self.irag = inteface_RAG
        self.kpath = ''
        self.reasoner = BaseReasoner

    def set_RAG(self, i_RAG): self.irag = i_RAG

    def set_path(self, path): self.kpath = path

    def set_reasoner(self, reasoner): self.reasoner = reasoner

    def getconcept(self, _concept):
        sysmsg = "You are an AI assistant."
        humanmsg = "Learn the professional terms. No output required."

        print(">>>>Searching the concept... " + _concept)
        problem = "What is " + _concept + " ?"
        self.irag.config({"KnowBase": self.kpath, "Questions": problem})
        ans = self.irag.answer()
        situations = {"txtHuman": humanmsg + '\n' + ans, "txtSystem": sysmsg}
        self.reasoner.set_text(situations)
        _results = self.reasoner.parser()
        return _results


def rec_expdata_to_excel(data_dict: dict, file_path: str, required_keys: list) -> None:
    for key in required_keys:
        if key not in data_dict:
            raise ValueError(f"miss some keywords: {key}")
    rdata = []
    for ikey in required_keys:
        rdata.append(data_dict.get(ikey))

    try:

        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "exp_data"

            headers = required_keys
            ws.append(headers)
        else:

            wb = load_workbook(file_path)
            ws = wb.active

        row_data = rdata
        ws.append(row_data)

        wb.save(file_path)
        print(f"sucessful writing: {file_path}")

    except PermissionError:
        raise PermissionError(f"can't write {file_path}，")
    except Exception as e:
        raise Exception(f"fail to write: {str(e)}")


def run_experiment(lstCase,turns,curpath):
    for turns in range(0,turns):  # set times of experiments, for example, range(0,1) is just one experiment, range(0,10) means ten times experiments
        print(f"-----The {turns}th set experiments-----")
        for cid in lstCase:
            knowcollector = inteface_RAG()
            objrag = clsRAG()
            knowcollector.RAG = objrag
            searcher = Searching()
            searcher.set_RAG(knowcollector)

            memoryofexp = MemoryManager()
            objexp = clsExperiments()
            objexp.expdata = DataExperiments()
            objexp.prompt = clsprompts()
            objexp.reasoner = BaseReasoner()
            objexp.k_collector = searcher

            objexp.reasoner.memory = memoryofexp

            objexp.dataid = cid
            objexp.advmaxtimes = 1
            objexp.sleep = 0
            objexp.experimentBuilder()

            fsp=os.path.join(curpath, "source")
            fname_expData = os.path.join(fsp, f"exp_data_case{cid}.xlsx")

            try:
                rec_expdata_to_excel(objexp.exp_dict, fname_expData, objexp.required_keys)
            except Exception as e:
                print(f"error: {e}")
            # ----------------release objects-------------------
            knowcollector = None
            objrag = None
            searcher = None
            memoryofexp = None
            objexp = None
            gc.collect()





